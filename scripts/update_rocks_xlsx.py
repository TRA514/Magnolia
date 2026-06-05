#!/usr/bin/env python3
"""Update the L10 Rocks tab with weekly Q2 metric values.

Locates the WAU and Stickiness rows by column-A label, finds the column
whose row-1 date matches --report-date, and writes the values. Optionally
inserts a new column if no match is found.

Companion to the /update-rocks command and the quarterly-rocks skill.
"""
from __future__ import annotations

import argparse
import datetime as dt
import sys
from copy import copy
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write(
        "ERROR: openpyxl is not installed.\n"
        "Install with: pip3 install -r requirements-rocks.txt\n"
    )
    sys.exit(2)


SHEET_NAME = "Rocks"
WAU_ROW_LABEL = "Home WAU (4-week rolling average)"
STICKINESS_ROW_LABEL = "Board Member Weekly Login Rate"
FIRST_DATA_COL = 5  # column E


def parse_date(s: str) -> dt.date:
    return dt.datetime.strptime(s, "%Y-%m-%d").date()


def cell_to_date(value) -> dt.date | None:
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    return None


def format_wau(wau: int) -> str:
    """Render WAU as e.g. '504k' or '504.6k'."""
    thousands = wau / 1000.0
    s = f"{thousands:.1f}"
    if s.endswith(".0"):
        s = s[:-2]
    return f"{s}k"


def find_label_row(ws, label: str) -> int:
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == label:
            return row
    raise SystemExit(f"ERROR: could not find row labeled {label!r} in column A.")


def collect_dated_columns(ws) -> list[tuple[int, dt.date]]:
    """Return [(col_idx, date)] for every row-1 cell from column E onward whose value is a date."""
    out: list[tuple[int, dt.date]] = []
    for col in range(FIRST_DATA_COL, ws.max_column + 1):
        d = cell_to_date(ws.cell(row=1, column=col).value)
        if d is not None:
            out.append((col, d))
    return out


def find_target_column(dated: list[tuple[int, dt.date]], report_date: dt.date) -> int | None:
    for col, d in dated:
        if d == report_date:
            return col
    return None


def insert_new_column(ws, dated: list[tuple[int, dt.date]], report_date: dt.date) -> int:
    """Append a new column to the right of the last dated column. Copy styles from prior column."""
    if not dated:
        new_col = FIRST_DATA_COL
        prior_col = None
    else:
        prior_col = max(c for c, _ in dated)
        new_col = prior_col + 1

    header_cell = ws.cell(row=1, column=new_col, value=report_date)
    if prior_col is not None:
        prior_header = ws.cell(row=1, column=prior_col)
        if prior_header.has_style:
            header_cell.font = copy(prior_header.font)
            header_cell.fill = copy(prior_header.fill)
            header_cell.border = copy(prior_header.border)
            header_cell.alignment = copy(prior_header.alignment)
            header_cell.number_format = prior_header.number_format
        else:
            header_cell.number_format = "m/d/yyyy"
        for row_idx in range(2, ws.max_row + 1):
            prior = ws.cell(row=row_idx, column=prior_col)
            new = ws.cell(row=row_idx, column=new_col)
            if prior.has_style:
                new.font = copy(prior.font)
                new.fill = copy(prior.fill)
                new.border = copy(prior.border)
                new.alignment = copy(prior.alignment)
                new.number_format = prior.number_format
    else:
        header_cell.number_format = "m/d/yyyy"
    return new_col


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", required=True, type=Path)
    ap.add_argument("--report-date", required=True, type=parse_date,
                    help="Column header date (typically a Thursday) in YYYY-MM-DD")
    ap.add_argument("--as-of-date", required=True, type=parse_date,
                    help="Saturday end-of-week the metrics were calculated for")
    ap.add_argument("--wau", required=True, type=int,
                    help="WAU value (integer count of unique visitors)")
    ap.add_argument("--stickiness", required=True, type=float,
                    help="Stickiness rate as a decimal (e.g. 0.3201 for 32.01%%)")
    ap.add_argument("--insert-if-missing", action="store_true",
                    help="Insert a new column if no header matches --report-date")
    ap.add_argument("--overwrite", action="store_true",
                    help="Overwrite cells even if they already have values")
    ap.add_argument("--dry-run", action="store_true",
                    help="Print planned changes without saving")
    args = ap.parse_args()

    if not args.workbook.exists():
        sys.stderr.write(f"ERROR: workbook not found: {args.workbook}\n")
        return 1

    if args.as_of_date.weekday() != 5:
        sys.stderr.write(
            f"WARN: --as-of-date {args.as_of_date} is a "
            f"{args.as_of_date.strftime('%A')}, not a Saturday. "
            "Continuing — quarterly-rocks skill convention is Saturday-anchored.\n"
        )

    try:
        wb_read = load_workbook(args.workbook, keep_vba=False, data_only=True)
        wb = load_workbook(args.workbook, keep_vba=False, data_only=False)
    except PermissionError:
        sys.stderr.write(
            f"ERROR: cannot open {args.workbook} (permission denied). "
            "Is the file open in Excel? Close it and retry.\n"
        )
        return 1

    if SHEET_NAME not in wb.sheetnames:
        sys.stderr.write(f"ERROR: workbook has no sheet named {SHEET_NAME!r}. "
                         f"Sheets present: {wb.sheetnames}\n")
        return 1
    ws_read = wb_read[SHEET_NAME]
    ws = wb[SHEET_NAME]

    wau_row = find_label_row(ws, WAU_ROW_LABEL)
    stick_row = find_label_row(ws, STICKINESS_ROW_LABEL)

    dated = collect_dated_columns(ws_read)
    target_col = find_target_column(dated, args.report_date)

    inserted = False
    if target_col is None:
        if not args.insert_if_missing:
            sys.stderr.write(
                f"ERROR: no column with header {args.report_date.isoformat()} found.\n"
                f"Existing dated columns: "
                f"{[(get_column_letter(c), d.isoformat()) for c, d in dated]}\n"
                f"Pass --insert-if-missing to append a new column.\n"
            )
            return 1
        target_col = insert_new_column(ws, dated, args.report_date)
        inserted = True

    target_letter = get_column_letter(target_col)
    wau_cell = ws.cell(row=wau_row, column=target_col)
    stick_cell = ws.cell(row=stick_row, column=target_col)

    existing_wau = wau_cell.value
    existing_stick = stick_cell.value
    if (existing_wau not in (None, "") or existing_stick not in (None, "")) and not args.overwrite and not inserted:
        sys.stderr.write(
            f"ERROR: column {target_letter} ({args.report_date.isoformat()}) "
            f"already has values:\n"
            f"  {WAU_ROW_LABEL} = {existing_wau!r}\n"
            f"  {STICKINESS_ROW_LABEL} = {existing_stick!r}\n"
            f"Pass --overwrite to replace.\n"
        )
        return 1

    new_wau_str = format_wau(args.wau)
    new_stick = round(args.stickiness, 4)

    print(f"Workbook : {args.workbook}")
    print(f"Sheet    : {SHEET_NAME}")
    print(f"Report   : {args.report_date.isoformat()} (column {target_letter}"
          + (" — NEW)" if inserted else ")"))
    print(f"As of    : {args.as_of_date.isoformat()}")
    print(f"WAU      : {target_letter}{wau_row}  {existing_wau!r:>16} -> {new_wau_str!r}")
    print(f"Sticky   : {target_letter}{stick_row}  {existing_stick!r:>16} -> {new_stick}")

    if args.dry_run:
        print("\nDRY RUN — no changes saved.")
        return 0

    wau_cell.value = new_wau_str
    stick_cell.value = new_stick

    try:
        wb.save(args.workbook)
    except PermissionError:
        sys.stderr.write(
            f"ERROR: cannot save {args.workbook} (permission denied). "
            "Is the file open in Excel? Close it and retry.\n"
        )
        return 1

    print("\nSaved.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
